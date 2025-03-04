import pandas as pd
import openpyxl
import numpy as np
from collections import defaultdict, Counter
import xlwings
from dotenv import load_dotenv

load_dotenv()


def detect_and_extract_tables(excel_file_path, sheet_name=None):
    """
    Detect and extract tables from an Excel sheet.

    Parameters:
    excel_file_path (str): Path to the Excel file.
    sheet_name (str, optional): Name of the sheet to analyze. If None, all sheets will be processed.

    Returns:
    list: A list of pandas DataFrames, each representing a detected table.
    """
    # Load the workbook
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(excel_file_path)
    excel_book.save()
    excel_book.close()
    excel_app.quit()

    # wb = openpyxl.load_workbook(excel_file_path, data_only=True)

    wb_data = openpyxl.load_workbook(excel_file_path, data_only=True)
    wb_formula = openpyxl.load_workbook(excel_file_path, data_only=False)

    # Get all sheets if sheet_name is None
    if sheet_name is None:
        sheets = wb_data.sheetnames
    else:
        sheets = [sheet_name]

    all_tables = []

    for sheet_name in sheets:
        sheet = wb_data[sheet_name]
        sheet_formula = wb_formula[sheet_name]
        # all_tables[sheet_name] = []

        # Get the data from the sheet and handle merged cells
        data = _get_data_with_merged_cells(sheet, sheet_formula)

        # Detect tables in the data
        tables = _detect_tables_in_data(data)

        for tbl in tables:
            tbl_, rn = set_next_row_as_header(tbl)
            if rn > 0:
                removed_tbl = tbl.head(rn - 1)
            else:
                removed_tbl = None

            all_tables.append(
                {"sheet_name": sheet_name, "table": tbl_, "removed": removed_tbl}
            )
            # all_tables[sheet_name].append(tbl_)

    return all_tables


def set_next_row_as_header(df):
    """Recursively sets the next row as the header if all column names are the same or contain NaN."""

    count_elements = Counter(df.columns)
    most_common_element, freq = count_elements.most_common(1)[0]
    rn = 0

    while (
        df.columns.nunique() / len(df.columns) < 0.2
        or freq / len(df.columns) > 0.5
        or df.columns.isnull().all()
    ):
        if len(df) > 1 and most_common_element != 0:  # Ensure there's a next row
            df.columns = df.iloc[0]  # Set next row as header
            df = df[1:]
            count_elements = Counter(df.columns)
            most_common_element, freq = count_elements.most_common(1)[0]
            rn = rn + 1
        else:
            break  # Stop if no more rows to process
    return df, rn


def _get_data_with_merged_cells(sheet_data, sheet_formula):
    """
    Get data from a sheet, handling merged cells.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): Excel sheet.

    Returns:
    numpy.ndarray: 2D array of cell values.
    """
    # Get the dimensions of the sheet
    max_row = sheet_data.max_row
    max_col = sheet_data.max_column

    # Initialize a 2D array to store the data
    data = np.empty((max_row, max_col), dtype=object)

    # Fill the array with cell values
    # First, get all values
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            # Get calculated value
            cell_value = sheet_data.cell(row=row, column=col).value

            # If value is None, check if it's a formula
            if cell_value is None:
                formula_cell = sheet_formula.cell(row=row, column=col)
                if formula_cell.data_type == "f":
                    # Store the formula text as fallback
                    cell_value = formula_cell.value

            data[row - 1, col - 1] = cell_value

    # Then handle merged cells
    for merged_range in sheet_data.merged_cells.ranges:
        # Get coordinates of the merged range
        min_row, min_col, max_r, max_c = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )

        # Get value from the top-left cell of the merged range
        top_left_value = data[min_row - 1, min_col - 1]

        # If top_left_value is None, try to get the formula
        if top_left_value is None:
            formula_cell = sheet_formula.cell(row=min_row, column=min_col)
            if formula_cell.data_type == "f":
                top_left_value = formula_cell.value

        # Fill the value into all cells in the merged range
        for r in range(min_row, max_r + 1):
            for c in range(min_col, max_c + 1):
                data[r - 1, c - 1] = top_left_value

    return data


def _detect_tables_in_data(data):
    """
    Detect and extract tables from a 2D array of data.

    Parameters:
    data (numpy.ndarray): 2D array of cell values.

    Returns:
    list: A list of pandas DataFrames, each representing a detected table.
    """
    # Create a binary mask indicating non-empty cells
    mask = np.vectorize(
        lambda x: x is not None and (not isinstance(x, str) or x.strip() != "")
    )(data)

    # Detect continuous rows with non-empty cells
    row_has_data = np.any(mask, axis=1)

    # Find table boundaries based on empty rows
    table_boundaries = []
    in_table = False
    start_idx = 0

    for i, has_data in enumerate(row_has_data):
        if has_data and not in_table:
            # Start of a new table
            start_idx = i
            in_table = True
        elif not has_data and in_table:
            # End of the current table
            table_boundaries.append((start_idx, i))
            in_table = False

    # Handle the case where the last table extends to the end of the data
    if in_table:
        table_boundaries.append((start_idx, len(row_has_data)))

    # If no table boundaries were found, return an empty list
    if not table_boundaries:
        return []

    # Analyze empty rows between tables to decide if tables should be merged
    merged_boundaries = [table_boundaries[0]]

    for i in range(1, len(table_boundaries)):
        prev_end = merged_boundaries[-1][1]
        curr_start = table_boundaries[i][0]

        in_between_empty_rows = curr_start - prev_end

        # If separated by exactly one empty row
        if in_between_empty_rows > 0 and in_between_empty_rows < 5:
            global rows_before, rows_after
            # Look at a few rows before and after the empty row
            rows_before = data[max(merged_boundaries[-1][0], prev_end - 8) : prev_end]
            rows_after = data[
                curr_start : min(
                    curr_start + 8 + in_between_empty_rows - 1, table_boundaries[i][1]
                )
            ]

            # Compare data patterns of rows before and after
            if _should_merge_tables(rows_before, rows_after):
                # Merge the tables
                merged_boundaries[-1] = (
                    merged_boundaries[-1][0],
                    table_boundaries[i][1],
                )
            else:
                # Keep as separate tables
                merged_boundaries.append(table_boundaries[i])
        else:
            # More than one empty row, always keep separate
            merged_boundaries.append(table_boundaries[i])

    # Extract tables from the merged boundaries
    tables = []
    for start, end in merged_boundaries:
        table_data = data[start:end]

        # Skip tables that are too small
        if (
            len(table_data) < 2
        ):  # Consider a table must have at least a header and a data row
            continue

        # Check if the first row is likely a header
        is_header = _is_likely_header(table_data[0], table_data[1:])

        # Create a pandas DataFrame
        if is_header:
            df = pd.DataFrame(table_data[1:], columns=table_data[0])
        else:
            df = pd.DataFrame(table_data)

        # Clean up the table by dropping empty columns
        df = df.dropna(axis=1, how="all")

        tables.append(df)

    return tables


def _should_merge_tables(rows_before, rows_after):
    # If either segment is empty, can't compare
    if len(rows_before) == 0 or len(rows_after) == 0:
        return False

    # 1. Compute general column pattern similarity (existing logic)
    non_empty_cols_before = []
    for row in rows_before:
        non_empty_cols_before.append(
            [
                i
                for i, cell in enumerate(row)
                if cell is not None
                and (not isinstance(cell, str) or cell.strip() != "")
            ]
        )

    non_empty_cols_after = []
    for row in rows_after:
        non_empty_cols_after.append(
            [
                i
                for i, cell in enumerate(row)
                if cell is not None
                and (not isinstance(cell, str) or cell.strip() != "")
            ]
        )

    # Flatten and get unique columns
    all_cols_before = set(
        [col for row_cols in non_empty_cols_before for col in row_cols]
    )
    all_cols_after = set([col for row_cols in non_empty_cols_after for col in row_cols])

    # Calculate column pattern similarity
    if not all_cols_before or not all_cols_after:
        return False

    col_similarity = len(all_cols_before & all_cols_after) / len(
        all_cols_before | all_cols_after
    )

    # 2. Check similarity of rows with minimum missing values
    missing_values_before = [
        sum(
            1
            for cell in row
            if cell is None or (isinstance(cell, str) and cell.strip() == "")
        )
        for row in rows_before
    ]
    missing_values_after = [
        sum(
            1
            for cell in row
            if cell is None or (isinstance(cell, str) and cell.strip() == "")
        )
        for row in rows_after
    ]

    if not missing_values_before or not missing_values_after:
        return False

    min_empty_before = min(missing_values_before)
    min_empty_after = min(missing_values_after)

    # Get rows with minimum missing values (these are likely the data rows, not headers or summaries)
    dense_rows_before = [
        row
        for i, row in enumerate(rows_before)
        if missing_values_before[i] == min_empty_before
    ]
    dense_rows_after = [
        row
        for i, row in enumerate(rows_after)
        if missing_values_after[i] == min_empty_after
    ]

    # Extract patterns of non-empty cells for dense rows
    dense_cols_before = []
    for row in dense_rows_before:
        dense_cols_before.append(
            [
                i
                for i, cell in enumerate(row)
                if cell is not None
                and (not isinstance(cell, str) or cell.strip() != "")
            ]
        )

    dense_cols_after = []
    for row in dense_rows_after:
        dense_cols_after.append(
            [
                i
                for i, cell in enumerate(row)
                if cell is not None
                and (not isinstance(cell, str) or cell.strip() != "")
            ]
        )

    # Flatten and get unique columns for dense rows
    all_dense_cols_before = set(
        [col for row_cols in dense_cols_before for col in row_cols]
    )
    all_dense_cols_after = set(
        [col for row_cols in dense_cols_after for col in row_cols]
    )

    # If no dense columns, fall back to general column similarity
    if not all_dense_cols_before or not all_dense_cols_after:
        return col_similarity > 0.6

    # Calculate similarity for dense rows
    dense_col_similarity = len(all_dense_cols_before & all_dense_cols_after) / len(
        all_dense_cols_before | all_dense_cols_after
    )

    # 3. Check if the first row of rows_after looks like a header (existing logic)
    first_row_after = rows_after[0]

    # Check if it's all strings and distinct from subsequent rows
    is_header_like = all(
        isinstance(cell, str) or cell is None for cell in first_row_after
    )

    if len(rows_after) > 1:
        # Compare data types with subsequent rows
        first_row_types = [type(cell) for cell in first_row_after if cell is not None]
        other_rows_types = [
            type(cell) for row in rows_after[1:] for cell in row if cell is not None
        ]

        if first_row_types and other_rows_types:
            first_row_type_counts = {
                t: first_row_types.count(t) / len(first_row_types)
                for t in set(first_row_types)
            }
            other_rows_type_counts = {
                t: other_rows_types.count(t) / len(other_rows_types)
                for t in set(other_rows_types)
            }

            # Check if type distributions are different
            type_diff = sum(
                abs(first_row_type_counts.get(t, 0) - other_rows_type_counts.get(t, 0))
                for t in set(first_row_type_counts.keys())
                | set(other_rows_type_counts.keys())
            )

            is_header_like = is_header_like and type_diff > 0.3

    # 4. Make the decision
    # Merge if either:
    # a) Dense rows have high similarity (this is the key improvement), or
    # b) Column patterns are similar and the first row isn't a header (from original logic)
    return dense_col_similarity > 0.7 or (col_similarity > 0.6 and not is_header_like)


def _is_likely_header(header_row, data_rows):
    """
    Check if a row is likely a header.

    Parameters:
    header_row (numpy.ndarray): Potential header row.
    data_rows (numpy.ndarray): Data rows.

    Returns:
    bool: True if the row is likely a header, False otherwise.
    """
    # Strategy 1: Check if the header row is all strings (common for headers)
    header_all_strings = all(
        isinstance(cell, str) or cell is None for cell in header_row
    )

    # header_all_valid = all(cell is None or isinstance(cell, (str, int, float)) for cell in header_row)

    # Strategy 2: Check if the header row has a different data type pattern than the data rows
    header_types = [type(cell) for cell in header_row if cell is not None]
    data_types = [type(cell) for row in data_rows for cell in row if cell is not None]

    header_type_counts = defaultdict(int)
    for t in header_types:
        header_type_counts[t] += 1

    data_type_counts = defaultdict(int)
    for t in data_types:
        data_type_counts[t] += 1

    # Normalize the counts
    total_header = sum(header_type_counts.values())
    total_data = sum(data_type_counts.values())

    type_diff = 0
    if total_header > 0 and total_data > 0:
        for t in set(header_type_counts.keys()) | set(data_type_counts.keys()):
            header_ratio = header_type_counts[t] / total_header
            data_ratio = (
                data_type_counts[t] / total_data if t in data_type_counts else 0
            )
            type_diff += abs(header_ratio - data_ratio)

    # Combine strategies
    return header_all_strings or type_diff > 0.5
